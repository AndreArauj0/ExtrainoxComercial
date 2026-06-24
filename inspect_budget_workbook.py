from __future__ import annotations

import json
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\andre\OneDrive\Área de Trabalho\Cópia de Desenvolvimento Orçamentos.xlsx")
OUT = Path("work/budget_workbook_profile.json")


def cell_payload(cell):
    value = cell.value
    if value is None:
        return None
    text = str(value)
    if len(text) > 160:
        text = text[:157] + "..."
    return {
        "address": cell.coordinate,
        "value": text,
        "data_type": cell.data_type,
        "style_id": cell.style_id,
        "number_format": cell.number_format,
    }


def inspect_sheet(ws, formula_ws=None):
    populated = []
    formulas = []
    max_row = min(ws.max_row, 120)
    max_col = min(ws.max_column, 40)

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            payload = cell_payload(cell)
            if payload:
                populated.append(payload)
            if formula_ws is not None:
                formula_value = formula_ws[cell.coordinate].value
                if isinstance(formula_value, str) and formula_value.startswith("="):
                    formulas.append(
                        {
                            "address": cell.coordinate,
                            "formula": formula_value[:220],
                            "cached_value": payload["value"] if payload else None,
                            "number_format": cell.number_format,
                        }
                    )

    row_density = []
    for idx in range(1, min(ws.max_row, 80) + 1):
        values = [
            ws.cell(idx, col).value
            for col in range(1, min(ws.max_column, 50) + 1)
            if ws.cell(idx, col).value is not None
        ]
        if values:
            row_density.append(
                {
                    "row": idx,
                    "filled": len(values),
                    "sample": [str(v)[:80] for v in values[:8]],
                }
            )

    column_headers = []
    for row in row_density[:30]:
        if row["filled"] >= 3:
            column_headers.append(row)

    return {
        "title": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "merged_ranges": [str(rng) for rng in list(ws.merged_cells.ranges)[:40]],
        "tables": list(ws.tables.keys()),
        "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter and ws.auto_filter.ref else None,
        "row_density": row_density[:40],
        "possible_header_rows": column_headers[:12],
        "sample_cells": populated[:260],
        "formulas": formulas[:220],
    }


def main():
    value_wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=False)
    formula_wb = openpyxl.load_workbook(SOURCE, data_only=False, read_only=False)

    profile = {
        "source": str(SOURCE),
        "sheets": [],
        "defined_names": [],
    }

    for name in value_wb.sheetnames:
        profile["sheets"].append(inspect_sheet(value_wb[name], formula_wb[name]))

    for defined_name in formula_wb.defined_names.values():
        destinations = []
        try:
            destinations = [f"{sheet}!{coord}" for sheet, coord in defined_name.destinations]
        except Exception:
            destinations = []
        profile["defined_names"].append(
            {
                "name": defined_name.name,
                "attr_text": defined_name.attr_text,
                "destinations": destinations,
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"sheets": value_wb.sheetnames, "output": str(OUT)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
