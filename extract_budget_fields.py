from __future__ import annotations

from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\andre\OneDrive\Área de Trabalho\Cópia de Desenvolvimento Orçamentos.xlsx")


def dump_row(value_ws, formula_ws, headers, row_index):
    print(f"\nROW {row_index}")
    for col_index, header in headers:
        cell = value_ws.cell(row_index, col_index)
        formula_cell = formula_ws.cell(row_index, col_index)
        print(
            f"{col_index:02d} {cell.coordinate} {header}: "
            f"value={cell.value!r} formula={formula_cell.value!r}"
        )


def main():
    values = openpyxl.load_workbook(SOURCE, data_only=True)
    formulas = openpyxl.load_workbook(SOURCE, data_only=False)
    value_ws = values["Orçamentos"]
    formula_ws = formulas["Orçamentos"]

    headers = [(col, value_ws.cell(3, col).value) for col in range(1, value_ws.max_column + 1)]

    print("HEADERS")
    for col, header in headers:
        print(f"{col:02d} {value_ws.cell(3, col).coordinate}: {header}")

    dump_row(value_ws, formula_ws, headers, 4)

    for row in range(4, value_ws.max_row + 1):
        if value_ws.cell(row, 1).value == "EXTRAINOX":
            dump_row(value_ws, formula_ws, headers, row)
            break

    print("\nCOMPANY SHEETS")
    for sheet_name in ["Extrafrio", "Extrainox"]:
        ws = values[sheet_name]
        fws = formulas[sheet_name]
        print(f"\n{sheet_name}")
        for row in range(1, 30):
            values_in_row = []
            for col in range(1, min(ws.max_column, 15) + 1):
                value = ws.cell(row, col).value
                formula = fws.cell(row, col).value
                if value is not None or formula is not None:
                    values_in_row.append(f"{ws.cell(row,col).coordinate}={formula!r}/{value!r}")
            if values_in_row:
                print(" | ".join(values_in_row))


if __name__ == "__main__":
    main()
